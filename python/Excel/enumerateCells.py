import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
rows = sheet.get_highest_row()
cols = sheet.get_highest_column()
for i in range(1, rows + 1):
	for j in range(1, cols + 1):
		print('%s: %s' % (sheet.cell(row=i, column=j).coordinate, sheet.cell(row=i, column=j).value))
	print('---------------------------------------------')

	