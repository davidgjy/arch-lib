import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

print(sheet.cell(row=1, column=2).value)
# from 1 to 8 step is 2
for i in range(1, sheet.max_row + 1, 2):
	print(i, sheet.cell(row=i, column=2).value)

print('------------------------------------------------')
	
# enumerate range cells
for rowOfCellObjects in sheet['A1':'C3']:
	for cellObj in rowOfCellObjects:
		print(cellObj.coordinate, cellObj.value)
	print('**************************************')

print('------------------------------------------------')

# enumerate the whole sheet
for i in range(1, sheet.max_row + 1):
	for j in range(1, sheet.max_column + 1):
		print(i, sheet.cell(row=i, column=j).coordinate, sheet.cell(row=i, column=j).value)
	print('**************************************')

	
	